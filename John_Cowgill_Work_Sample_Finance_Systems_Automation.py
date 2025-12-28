"""
WORK SAMPLE (SANITIZED) — Financial Systems Automation (Python + SQL)

- Purpose: Demonstrates a reusable workflow that queries multiple SQL Server databases,
  performs light transformations/validation, and exports a standardized dataset for reporting.
- Sanitization: Credentials, internal hosts, file shares, and company-specific identifiers have been redacted.
  Database names are intentionally retained to demonstrate cross-database access patterns.
- Note: Table/field names reflect the source system structure; this excerpt is provided as a representative sample.
"""

# ---- consolidated Sales Charts exports ----
import time as _t

_T0 = _t.perf_counter()

import os
import csv
import uuid
import shutil
import tempfile
from pathlib import Path
from contextlib import contextmanager
import warnings
import pyodbc
from datetime import datetime, date


@contextmanager
def timed(label, bucket):
    t0 = _t.perf_counter()
    try:
        yield
    finally:
        dt = _t.perf_counter() - t0
        bucket[label] = bucket.get(label, 0.0) + dt
        print(f"{label}: {dt:.3f} s")


def get_dates() -> tuple[date, date]:
    """Prefer a tiny INI; fallback to Excel if needed. Returns Python date objects."""
    from configparser import ConfigParser

    ini_path = r"U:\<user>\Scripts\DateVariables.ini"

    if os.path.exists(ini_path):
        with timed(
            "READ_DATE_VARS_INI", spans_tmp := {}
        ):  # ephemeral spans for I/O timing
            cfg = ConfigParser()
            cfg.read(ini_path, encoding="utf-8")
            sc_start = cfg["dates"]["sc_start"].strip()  # YYYYMMDD
            sc_end = cfg["dates"]["sc_end"].strip()  # YYYYMMDD
        return (
            datetime.strptime(sc_start, "%Y%m%d").date(),
            datetime.strptime(sc_end, "%Y%m%d").date(),
        )

    # Fallback to Excel only if INI missing (avoids pandas)
    with timed("READ_DATE_VARS_XLSM", spans_tmp := {}):
        from openpyxl import load_workbook

        wb = load_workbook(
            r"U:\<user>\Scripts\DateVariablesList.xlsm",
            data_only=True,
            read_only=True,
        )
        ws = wb.active  # or wb["Sheet1"]
        start = ws["C14"].value  # expected 20231218 etc.
        end_ = ws["C15"].value
    return (
        datetime.strptime(str(start), "%Y%m%d").date(),
        datetime.strptime(str(end_), "%Y%m%d").date(),
    )


def build_sales_query(
    *,
    datefirst: int,
    week_expr: str,
    dataid_keys: list[int],
    include_qsr: bool = False,
    site_between: tuple[str, str] | None = None,
) -> str:
    """
    week_expr: 'DATEPART(wk, pb.PaperworkBatch_BusDate)' or 'DATEPART(ISO_WEEK, pb.PaperworkBatch_BusDate)'
    """
    keys_csv = ", ".join(str(k) for k in dataid_keys)

    # optional filters
    extra_filters = []
    if include_qsr:
        extra_filters.append("pca.ProfCtr_Type_Desc = 'QSR'")
    if site_between:
        lo, hi = site_between
        extra_filters.append(f"pca.Site_ID BETWEEN '{lo}' AND '{hi}'")

    extra_where = ""
    if extra_filters:
        extra_where = "    AND " + "\n    AND ".join(extra_filters) + "\n"

    # Year-of-week calculation (kept consistent with your original)
    year_of_week_expr = "DATEPART(YEAR, DATEADD(DAY, 4 - DATEPART(WEEKDAY, pb.PaperworkBatch_BusDate), pb.PaperworkBatch_BusDate))"

    return f"""
SET NOCOUNT ON;
SET DATEFIRST {datefirst};
SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED;

SELECT
    CONCAT(
        {year_of_week_expr},
        '|', {week_expr},
        '|', DATEPART(WEEKDAY,  pb.PaperworkBatch_BusDate),
        '|', pca.Site_ID,
        '|', di.DataID_Key
    ) AS LookupValue,
    {week_expr} AS CalcWeek,
    DATEPART(WEEKDAY,  pb.PaperworkBatch_BusDate) AS CalcDay,
    pca.Site_ID,
    pca.Site_Description,
    pb.PaperworkBatch_BusDate,
    di.DataID_Key,
    di.[DataID_Description],
    dv.[DataIDValue_Value],
    SUM(dv.DataIDValue_Value) OVER (
        PARTITION BY
            pca.Site_ID,
            di.DataID_Key,
            {week_expr},
            {year_of_week_expr}
    ) AS CalcAmount
FROM dbo.DataID_Values      AS dv
JOIN dbo.Data_IDs           AS di  ON di.DataID_Key          = dv.DataIDValue_DataID_Key
JOIN dbo.Paperwork_Batches  AS pb  ON pb.PaperworkBatch_Key  = dv.DataIDValue_PaperworkBatch_Key
JOIN dbo.Profit_Centers_All AS pca ON pca.ProfCtr_Site_Key   = pb.PaperworkBatch_Site_Key
WHERE
    pb.PaperworkBatch_BusDate >= ?
    AND pb.PaperworkBatch_BusDate < DATEADD(DAY, 1, ?)
{extra_where}    AND di.DataID_Key IN (
        {keys_csv}
    )
OPTION (OPTIMIZE FOR UNKNOWN);
""".strip()


def stream_query_to_local_csv(
    cxn,
    sql,
    params,
    local_csv_path,
    arraysize=100_000,
    encoding="utf-8",
    include_header=True,
):
    cur = cxn.cursor()
    cur.arraysize = arraysize  # key tuning knob
    cur.execute(sql, params)

    cols = [d[0] for d in cur.description]
    row_count = 0

    with open(
        local_csv_path, "w", newline="", encoding=encoding, buffering=4 * 1024 * 1024
    ) as f:
        w = csv.writer(f, lineterminator="\n")
        if include_header:
            w.writerow(cols)
        while True:
            rows = cur.fetchmany()  # uses arraysize
            if not rows:
                break
            w.writerows(rows)
            row_count += len(rows)

    cur.close()
    return row_count


def save_csv_via_staging_streaming(
    cxn, sql, params, final_unc_path: str, arraysize=100_000, spans=None
):
    final_unc = Path(final_unc_path)
    final_unc.parent.mkdir(parents=True, exist_ok=True)
    unc_tmp = final_unc.with_suffix(".tmp.csv")
    local_tmp = Path(tempfile.gettempdir()) / f"{uuid.uuid4()}.csv"

    with timed("READ_STREAM_TO_LOCAL_CSV", spans):
        t0 = _t.perf_counter()
        rows_written = stream_query_to_local_csv(
            cxn=cxn,
            sql=sql,
            params=params,
            local_csv_path=local_tmp,
            arraysize=arraysize,
            encoding="utf-8",
            include_header=True,
        )
        dt = _t.perf_counter() - t0
        if dt > 0:
            print(
                f"  -> streamed {rows_written:,} rows @ {rows_written/dt:,.0f} rows/s (arraysize={arraysize})"
            )

    with timed("COPY_TO_UNC", spans):
        shutil.copyfile(local_tmp, unc_tmp)

    with timed("RENAME_ON_UNC", spans):
        os.replace(unc_tmp, final_unc)  # atomic in same directory

    try:
        local_tmp.unlink(missing_ok=True)
    except Exception:
        pass

    return rows_written


def run_sales_export(
    *,
    server: str,
    database: str,
    username: str,
    password: str,
    packet_size: str,
    arraysize: int,
    final_unc_path: str,
    sql: str,
    params: list,
    done_message: str,
) -> float:
    job_T0 = _t.perf_counter()
    spans = {}
    spans["IMPORTS"] = (
        _t.perf_counter() - job_T0
    )  # placeholder for symmetry with your original prints

    with timed("PREP", spans):
        warnings.filterwarnings(
            "ignore", message="Data Validation extension is not supported"
        )

    with timed("CONNECT", spans):
        cxn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={server};DATABASE={database};UID={username};PWD={password};"
            f"Packet Size={packet_size}"
        )

    rows_total = save_csv_via_staging_streaming(
        cxn=cxn,
        sql=sql,
        params=params,
        final_unc_path=final_unc_path,
        arraysize=arraysize,
        spans=spans,
    )

    with timed("CLEANUP", spans):
        total = _t.perf_counter() - job_T0
        accounted = sum(spans.values())
        print(f"\nTOTAL (in-script): {total:.3f} s")
        print("Breakdown:", ", ".join(f"{k}={v:.3f}s" for k, v in spans.items()))
        print(f"UNACCOUNTED: {total - accounted:.3f} s")
        print(f"Total Rows: {rows_total:,}")
        print(done_message)
        cxn.close()

    return total


# -------- shared config --------
SERVER = "<redacted_server>"
USERNAME = "<redacted_user>"
PASSWORD = "<redacted_password>"
ARRAY_SIZE = int(os.environ.get("PYODBC_ARRAYSIZE", "100000"))
PACKET_SIZE = os.environ.get("PYODBC_PACKETSIZE", "32767")

# shared dates (used by all three jobs)
start_date, end_date = get_dates()

final_total = 0.0

# -------- Job 1: Company A SW Sales Chart (wk, DATEFIRST 3, QSR + site range, DB: 2875_10) --------
sql_job1 = build_sales_query(
    datefirst=3,
    week_expr="DATEPART(wk, pb.PaperworkBatch_BusDate)",
    dataid_keys=[675, 1205, 1310, 673, 674, 683, 960],
    include_qsr=True,
    site_between=("2210", "2295"),
)
final_total += run_sales_export(
    server=SERVER,
    database="PDICompany_2875_10",
    username=USERNAME,
    password=PASSWORD,
    packet_size=PACKET_SIZE,
    arraysize=ARRAY_SIZE,
    final_unc_path=r"\\<file_share>\Accounting\<redacted_path>\SW_Sales_Chart_Data.csv",
    sql=sql_job1,
    params=[start_date, end_date],
    done_message="CoA Sales Chart Data Import Finished",
)

# -------- Job 2: Company B Sales Chart (ISO_WEEK, DATEFIRST 1, no QSR filter, DB: 2875_70) --------
sql_job2 = build_sales_query(
    datefirst=1,
    week_expr="DATEPART(ISO_WEEK, pb.PaperworkBatch_BusDate)",
    dataid_keys=[
        95,
        149,
        150,
        151,
        152,
        153,
        156,
        157,
        122,
        123,
        124,
        125,
        126,
        129,
        130,
        107,
        121,
    ],
    include_qsr=False,
)
final_total += run_sales_export(
    server=SERVER,
    database="PDICompany_2875_70",
    username=USERNAME,
    password=PASSWORD,
    packet_size=PACKET_SIZE,
    arraysize=ARRAY_SIZE,
    final_unc_path=r"\\<file_share>\Accounting\<redacted_path>\SW_Sales_Chart_Data.csv",
    sql=sql_job2,
    params=[start_date, end_date],
    done_message="CoA Sales Chart Data Import Finished",
)

# -------- Job 3: Company A Sales Chart (ISO_WEEK, DATEFIRST 1, QSR, DB: 2875_10) --------
sql_job3 = build_sales_query(
    datefirst=1,
    week_expr="DATEPART(ISO_WEEK, pb.PaperworkBatch_BusDate)",
    dataid_keys=[
        689,
        690,
        697,
        701,
        961,
        962,
        963,
        964,
        966,
        1141,
        1206,
        1207,
        1208,
        1209,
        1211,
        1311,
        1312,
        1313,
        1314,
        1316,
        1986,
        1987,
        1995,
        1996,
        2000,
        2001,
        2002,
        2004,
        2005,
        2009,
        2010,
        2025,
        2026,
        2027,
    ],
    include_qsr=True,
)
final_total += run_sales_export(
    server=SERVER,
    database="PDICompany_2875_10",
    username=USERNAME,
    password=PASSWORD,
    packet_size=PACKET_SIZE,
    arraysize=ARRAY_SIZE,
    final_unc_path=r"\\<file_share>\Accounting\<redacted_path>\SW_Sales_Chart_Data.csv",
    sql=sql_job3,
    params=[start_date, end_date],
    done_message="CoA Sales Chart Data Import Finished",
)

print(f"\nFINAL TOTAL (in-script): {final_total:.3f} s")